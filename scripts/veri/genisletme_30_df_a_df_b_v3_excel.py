"""
GENIŞLETME AŞAMA 30 — DF-A v3 ve DF-B v3'ün okunabilir Excel (.xlsx)
haline aktarilmasi.

Proje sahibinin isteği uzerine: 29 numarali gorevde kurulan
df_a_v3_noter_penceresi_2015_bugun.csv ve df_b_v3_enag_betam_2024_bugun.csv
dosyalarini Excel'de gozle kolayca inceleyebilmek icin. Veriyi DEGISTIRMEZ,
27 numarali gorevdeki AYNI bicimlendirme kurallarini (bold baslik,
dondurulmus baslik/tarih sutunu, otomatik filtre, renk kodlu sutun
gruplari, tarih/sayi bicimleri) iki yeni dosyaya uygular.

Girdi: data/processed/dataframes/df_a_v3_noter_penceresi_2015_bugun.csv
       data/processed/dataframes/df_b_v3_enag_betam_2024_bugun.csv
Cikti: data/processed/dataframes/df_a_v3_noter_penceresi_2015_bugun.xlsx
       data/processed/dataframes/df_b_v3_enag_betam_2024_bugun.xlsx
"""
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_KOKU = Path(__file__).resolve().parents[2]
DF_DIR = REPO_KOKU / "data" / "processed" / "dataframes"

REFERANS_AY_RENK = "FFF2CC"  # acik sari - "...referans_ay" sutunlari
GUNLUK_KUR_RENK = "D9E1F2"  # acik mavi - usdtry/eurtry
TAKVIM_RENK = "E2EFDA"  # acik yesil - yil/ay/gun/ceyrek/haftanin_gunu/yilin_gunu
DF_B_OZGU_RENK = "FCE4D6"  # acik turuncu - yalnizca DF-B'de olan (enag/proxy/noter_otomobil)

DF_A_OZGU_OLMAYAN = {
    "noter_devir_otomobil_adet",
    "enag_referans_ay", "enag_aylik_degisim", "enag_yillik_degisim",
    "proxy_referans_ay", "proxy_fiyat_cari_tl", "proxy_dom_gun", "proxy_satis_orani_pct",
    "alim_gucu_referans_ay", "brut_ucret_maas_endeksi_2021_100",
}


def bicimlendirilmis_yaz(kaynak_csv: Path, hedef_xlsx: Path, sayfa_adi: str):
    df = pd.read_csv(kaynak_csv, parse_dates=["tarih"])

    with pd.ExcelWriter(hedef_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sayfa_adi)
        ws = writer.sheets[sayfa_adi]

        baslik_yazi_tipi = Font(bold=True, color="FFFFFF")
        baslik_dolgu = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for hucre in ws[1]:
            hucre.font = baslik_yazi_tipi
            hucre.fill = baslik_dolgu
            hucre.alignment = Alignment(horizontal="center", vertical="center")

        for idx, kolon in enumerate(df.columns, start=1):
            harf = get_column_letter(idx)
            if kolon == "tarih":
                genislik = 12
                for hucre in ws[harf][1:]:
                    hucre.number_format = "yyyy-mm-dd"
            elif kolon.endswith("_referans_ay"):
                genislik = max(14, len(kolon) + 2)
                renk = DF_B_OZGU_RENK if kolon in DF_A_OZGU_OLMAYAN else REFERANS_AY_RENK
                for hucre in ws[harf][1:]:
                    hucre.fill = PatternFill(start_color=renk, end_color=renk, fill_type="solid")
            elif kolon.startswith("usdtry_") or kolon.startswith("eurtry_"):
                genislik = max(12, len(kolon) + 2)
                for hucre in ws[harf][1:]:
                    hucre.fill = PatternFill(start_color=GUNLUK_KUR_RENK, end_color=GUNLUK_KUR_RENK, fill_type="solid")
                    hucre.number_format = "0.0000"
            elif kolon in {"yil", "ay", "gun", "ceyrek", "haftanin_gunu", "yilin_gunu"}:
                genislik = max(10, len(kolon) + 2)
                for hucre in ws[harf][1:]:
                    hucre.fill = PatternFill(start_color=TAKVIM_RENK, end_color=TAKVIM_RENK, fill_type="solid")
            elif kolon in DF_A_OZGU_OLMAYAN:
                genislik = max(14, len(kolon) + 2)
                for hucre in ws[harf][1:]:
                    hucre.fill = PatternFill(start_color=DF_B_OZGU_RENK, end_color=DF_B_OZGU_RENK, fill_type="solid")
                    if pd.api.types.is_numeric_dtype(df[kolon]):
                        hucre.number_format = "#,##0.00"
            else:
                genislik = max(14, len(kolon) + 2)
                if pd.api.types.is_numeric_dtype(df[kolon]):
                    for hucre in ws[harf][1:]:
                        hucre.number_format = "#,##0.00"
            ws.column_dimensions[harf].width = min(genislik, 32)

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

    print(f"Yazildi: {hedef_xlsx} ({df.shape[0]} satir x {df.shape[1]} sutun)")


def main():
    bicimlendirilmis_yaz(
        DF_DIR / "df_a_v3_noter_penceresi_2015_bugun.csv",
        DF_DIR / "df_a_v3_noter_penceresi_2015_bugun.xlsx",
        "df_a_v3",
    )
    bicimlendirilmis_yaz(
        DF_DIR / "df_b_v3_enag_betam_2024_bugun.csv",
        DF_DIR / "df_b_v3_enag_betam_2024_bugun.xlsx",
        "df_b_v3",
    )


if __name__ == "__main__":
    main()
