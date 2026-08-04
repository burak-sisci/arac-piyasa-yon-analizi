"""
GENIŞLETME AŞAMA 37 — Korelasyon analizi fazının 4 güncel veri setini
(DF-A, DF-B, DF-A-log, DF-B-log) okunabilir Excel (.xlsx) haline aktarir.

(Korelasyon analizi fazı, proje sahibinin adım-adım talimatıyla)

27/30 nolu gorevlerdeki AYNI bicimlendirme kurallari (bold baslik,
dondurulmus baslik/tarih sutunu, otomatik filtre, sayi/tarih bicimleri,
_log_degisim sutunlari icin ayri renk) uygulanir. Veri DEGISTIRILMEZ.

Girdi/Cikti (ayni klasorde, .csv yaninda .xlsx):
  data/processed/dataframes/df_a_v3_noter_penceresi_2015_bugun.(csv|xlsx)
  data/processed/dataframes/df_b_v3_enag_betam_2024_bugun.(csv|xlsx)
  data/processed/dataframes/df_a_log_degisim_2015_bugun.(csv|xlsx)
  data/processed/dataframes/df_b_log_degisim_2024_bugun.(csv|xlsx)
"""
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_KOKU = Path(__file__).resolve().parents[2]
DF_DIR = REPO_KOKU / "data" / "processed" / "dataframes"

LOG_DEGISIM_RENK = "FCE4D6"  # acik turuncu - "..._log_degisim" sutunlari
TARGET_RENK = "C6E0B4"  # acik yesil - target sutunu

DOSYALAR = [
    (DF_DIR / "df_a_v3_noter_penceresi_2015_bugun.csv", DF_DIR / "df_a_v3_noter_penceresi_2015_bugun.xlsx", "df_a_v3", "noter_devir_toplam_adet"),
    (DF_DIR / "df_b_v3_enag_betam_2024_bugun.csv", DF_DIR / "df_b_v3_enag_betam_2024_bugun.xlsx", "df_b_v3", "noter_devir_toplam_adet"),
    (DF_DIR / "df_a_log_degisim_2015_bugun.csv", DF_DIR / "df_a_log_degisim_2015_bugun.xlsx", "df_a_log", "noter_devir_toplam_adet_log_degisim"),
    (DF_DIR / "df_b_log_degisim_2024_bugun.csv", DF_DIR / "df_b_log_degisim_2024_bugun.xlsx", "df_b_log", "noter_devir_toplam_adet_log_degisim"),
]


def main():
    for kaynak_csv, hedef_xlsx, sayfa_adi, target_kolon in DOSYALAR:
        df = pd.read_csv(kaynak_csv, parse_dates=["tarih"])

        with pd.ExcelWriter(hedef_xlsx, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sayfa_adi)
            ws = writer.sheets[sayfa_adi]

            baslik_yazi_tipi = Font(bold=True, color="FFFFFF")
            baslik_dolgu = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for hucre in ws[1]:
                hucre.font = baslik_yazi_tipi
                hucre.fill = baslik_dolgu
                hucre.alignment = Alignment(horizontal="center", vertical="center")

            for idx, kolon in enumerate(df.columns, start=1):
                harf = get_column_letter(idx)
                if kolon == "tarih":
                    genislik = 12
                    for hucre in ws[harf][1:]:
                        hucre.number_format = "yyyy-mm-dd"
                elif kolon == target_kolon:
                    genislik = max(20, len(kolon) + 2)
                    for hucre in ws[harf][1:]:
                        hucre.fill = PatternFill(start_color=TARGET_RENK, end_color=TARGET_RENK, fill_type="solid")
                        hucre.number_format = "#,##0.0000" if kolon.endswith("_log_degisim") else "#,##0.00"
                elif kolon.endswith("_log_degisim"):
                    genislik = max(16, len(kolon) + 2)
                    for hucre in ws[harf][1:]:
                        hucre.fill = PatternFill(start_color=LOG_DEGISIM_RENK, end_color=LOG_DEGISIM_RENK, fill_type="solid")
                        hucre.number_format = "0.0000"
                else:
                    genislik = max(14, len(kolon) + 2)
                    if pd.api.types.is_numeric_dtype(df[kolon]):
                        for hucre in ws[harf][1:]:
                            hucre.number_format = "#,##0.00"
                ws.column_dimensions[harf].width = min(genislik, 32)

            ws.freeze_panes = "B2"
            ws.auto_filter.ref = ws.dimensions

        print(f"Yazildi: {hedef_xlsx} ({df.shape[0]} satir x {df.shape[1]} sutun)")


if __name__ == "__main__":
    main()
