"""
GENIŞLETME AŞAMA 27 — df_gunluk_forward_fill_2015_bugun.csv'nin okunabilir
Excel (.xlsx) haline aktarilmasi.

Proje sahibinin isteği uzerine: gorev 26'nin ay-hizali doldurulmus gunluk
tablosunu Excel'de gozle kolayca inceleyebilmek icin. Veriyi DEGISTIRMEZ,
yalnizca ayni veriyi bicimlendirilmis (bold baslik, dondurulmus baslik
satiri/tarih sutunu, otomatik filtre, sutun genislikleri, tarih/sayi
bicimleri) bir .xlsx dosyasina yazar.

Girdi: data/processed/dataframes/df_gunluk_forward_fill_2015_bugun.csv
Cikti: data/processed/dataframes/df_gunluk_forward_fill_2015_bugun.xlsx
"""
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_KOKU = Path(__file__).resolve().parents[2]
DF_DIR = REPO_KOKU / "data" / "processed" / "dataframes"
KAYNAK_CSV = DF_DIR / "df_gunluk_forward_fill_2015_bugun.csv"
HEDEF_XLSX = DF_DIR / "df_gunluk_forward_fill_2015_bugun.xlsx"

REFERANS_AY_RENK = "FFF2CC"  # acik sari - "...referans_ay" sutunlari
GUNLUK_KUR_RENK = "D9E1F2"  # acik mavi - usdtry/eurtry
TAKVIM_RENK = "E2EFDA"  # acik yesil - yil/ay/gun/ceyrek/haftanin_gunu/yilin_gunu


def main():
    df = pd.read_csv(KAYNAK_CSV, parse_dates=["tarih"])

    with pd.ExcelWriter(HEDEF_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="gunluk_ay_hizali")
        ws = writer.sheets["gunluk_ay_hizali"]

        baslik_yazi_tipi = Font(bold=True, color="FFFFFF")
        baslik_dolgu = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for hucre in ws[1]:
            hucre.font = baslik_yazi_tipi
            hucre.fill = baslik_dolgu
            hucre.alignment = Alignment(horizontal="center", vertical="center")

        for idx, kolon in enumerate(df.columns, start=1):
            harf = get_column_letter(idx)
            if kolon == "tarih":
                genislik = 12
                for hucre in ws[harf][1:]:
                    hucre.number_format = "yyyy-mm-dd"
            elif kolon.endswith("_referans_ay"):
                genislik = max(14, len(kolon) + 2)
                for hucre in ws[harf][1:]:
                    hucre.fill = PatternFill(start_color=REFERANS_AY_RENK, end_color=REFERANS_AY_RENK, fill_type="solid")
            elif kolon.startswith("usdtry_") or kolon.startswith("eurtry_"):
                genislik = max(12, len(kolon) + 2)
                for hucre in ws[harf][1:]:
                    hucre.fill = PatternFill(start_color=GUNLUK_KUR_RENK, end_color=GUNLUK_KUR_RENK, fill_type="solid")
                    hucre.number_format = "0.0000"
            elif kolon in {"yil", "ay", "gun", "ceyrek", "haftanin_gunu", "yilin_gunu"}:
                genislik = max(10, len(kolon) + 2)
                for hucre in ws[harf][1:]:
                    hucre.fill = PatternFill(start_color=TAKVIM_RENK, end_color=TAKVIM_RENK, fill_type="solid")
            else:
                genislik = max(14, len(kolon) + 2)
                if pd.api.types.is_numeric_dtype(df[kolon]):
                    for hucre in ws[harf][1:]:
                        hucre.number_format = "#,##0.00"
            ws.column_dimensions[harf].width = min(genislik, 32)

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

    print(f"Yazildi: {HEDEF_XLSX}")
    print(f"Boyut: {df.shape[0]} satir x {df.shape[1]} sutun")


if __name__ == "__main__":
    main()
