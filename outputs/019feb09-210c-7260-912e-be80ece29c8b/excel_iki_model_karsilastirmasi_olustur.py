from pathlib import Path

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "outputs" / "autogluon" / "ip7_devir_orani" / "test_tahminleri_tum_modeller.csv"
OUTPUT = Path(__file__).with_name("iki_model_tahmin_karsilastirmasi.xlsx")
PREVIEW = Path(__file__).with_name("iki_model_tahmin_karsilastirmasi_onizleme.png")

MONTHS_TR = {
    1: "Ocak",
    2: "Şubat",
    3: "Mart",
    4: "Nisan",
    5: "Mayıs",
    6: "Haziran",
}

# 24 ay validasyon / 12 ay test düzeninde validasyonla seçilen Toto2'nin
# daha önce kaydedilmiş Ocak-Haziran 2026 tahminleri.
OLD_TOTO2 = {
    "2026-01-01": 0.042230669409036636,
    "2026-02-01": 0.034867238253355026,
    "2026-03-01": 0.03400428965687752,
    "2026-04-01": 0.03485333174467087,
    "2026-05-01": 0.03539777174592018,
    "2026-06-01": 0.03262593597173691,
}

raw = pd.read_csv(SOURCE, parse_dates=["hedef_ay"])
new_model = raw.loc[
    raw["model"].eq("RecursiveTabular"), ["hedef_ay", "y_true", "y_pred"]
].rename(columns={"y_pred": "model_6_6"})
last_year = raw.loc[
    raw["model"].eq("SeasonalNaive"), ["hedef_ay", "y_pred"]
].rename(columns={"y_pred": "gecen_yil"})

comparison = new_model.merge(last_year, on="hedef_ay", validate="one_to_one")
comparison = comparison.loc[comparison["hedef_ay"].dt.year.eq(2026)].sort_values("hedef_ay")
comparison["model_24_12"] = comparison["hedef_ay"].dt.strftime("%Y-%m-%d").map(OLD_TOTO2)

if len(comparison) != 6 or comparison.isna().any().any():
    raise ValueError("İki model için altı ortak ve eksiksiz test ayı bulunamadı.")

wb = Workbook()
ws = wb.active
ws.title = "Tahmin Karşılaştırması"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:F1")
ws["A1"] = "İki Eğitim Düzeninin Tahmin Karşılaştırması"
ws["A1"].font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="17365D")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

headers = [
    "Tahmin edilen ay",
    "Yıl",
    "12 ay test 24 ay validasyon modeli",
    "6 ay validasyon 6 ay test modeli",
    "Gerçek tahmin",
    "Geçen yılın o ayki modeli",
]

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F75B5")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, row in enumerate(comparison.itertuples(index=False), start=4):
    ws.cell(row=row_idx, column=1, value=MONTHS_TR[row.hedef_ay.month])
    ws.cell(row=row_idx, column=2, value=int(row.hedef_ay.year))
    ws.cell(row=row_idx, column=3, value=float(row.model_24_12))
    ws.cell(row=row_idx, column=4, value=float(row.model_6_6))
    ws.cell(row=row_idx, column=5, value=float(row.y_true))
    ws.cell(row=row_idx, column=6, value=float(row.gecen_yil))

    for col in range(1, 7):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(name="Aptos", size=11)
        cell.alignment = Alignment(
            horizontal="left" if col == 1 else "right",
            vertical="center",
        )
    for col in range(3, 7):
        ws.cell(row=row_idx, column=col).number_format = "0.000%"

last_data_row = 3 + len(comparison)
table = Table(displayName="IkiModelTahminKarsilastirma", ref=f"A3:F{last_data_row}")
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws.add_table(table)

ws.column_dimensions["A"].width = 21
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 34
ws.column_dimensions["D"].width = 33
ws.column_dimensions["E"].width = 19
ws.column_dimensions["F"].width = 29
ws.row_dimensions[3].height = 48
for row_idx in range(4, last_data_row + 1):
    ws.row_dimensions[row_idx].height = 21

note_row = last_data_row + 2
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
ws.cell(
    note_row,
    1,
    "Not: 24/12 modeli = Toto2; 6/6 modeli = RecursiveTabular; geçen-yıl modeli = aynı ayın 12 ay önceki gerçekleşmesi.",
)
ws.cell(note_row, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[note_row].height = 28

ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:F{last_data_row}"
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.print_title_rows = "1:3"
ws.print_area = f"A1:F{note_row}"

wb.save(OUTPUT)

# Yeniden açarak sütunlar, satırlar, değer türleri ve yüzde biçimini doğrula.
check = load_workbook(OUTPUT, data_only=False)
check_ws = check["Tahmin Karşılaştırması"]
assert [check_ws.cell(3, c).value for c in range(1, 7)] == headers
assert check_ws.max_row == note_row
assert all(
    isinstance(check_ws.cell(r, c).value, (int, float))
    for r in range(4, last_data_row + 1)
    for c in range(2, 7)
)
assert all(
    check_ws.cell(r, c).number_format == "0.000%"
    for r in range(4, last_data_row + 1)
    for c in range(3, 7)
)

# Görsel düzen kontrolü için çalışma sayfasının kompakt bir önizlemesini üret.
preview_rows = []
for r in range(4, last_data_row + 1):
    preview_rows.append(
        [
            check_ws.cell(r, 1).value,
            str(check_ws.cell(r, 2).value),
            *[f"%{100 * check_ws.cell(r, c).value:.3f}" for c in range(3, 7)],
        ]
    )

plt.rcParams["font.family"] = "DejaVu Sans"
fig, ax = plt.subplots(figsize=(17, 4.8), facecolor="white")
ax.axis("off")
fig.suptitle("İki Eğitim Düzeninin Tahmin Karşılaştırması", fontsize=18, fontweight="bold", y=0.96)
preview_table = ax.table(
    cellText=preview_rows,
    colLabels=headers,
    cellLoc="center",
    colWidths=[0.12, 0.07, 0.24, 0.23, 0.14, 0.20],
    bbox=[0.01, 0.14, 0.98, 0.68],
)
preview_table.auto_set_font_size(False)
preview_table.set_fontsize(10)
for (r, c), cell in preview_table.get_celld().items():
    cell.set_edgecolor("#D9E2F3")
    if r == 0:
        cell.set_facecolor("#2F75B5")
        cell.get_text().set_color("white")
        cell.get_text().set_fontweight("bold")
        cell.get_text().set_wrap(True)
    elif r % 2:
        cell.set_facecolor("#EAF2F8")
fig.text(
    0.5,
    0.06,
    "24/12 modeli: Toto2 | 6/6 modeli: RecursiveTabular | Ortak dönem: Ocak–Haziran 2026",
    ha="center",
    fontsize=10,
    color="#666666",
)
fig.savefig(PREVIEW, dpi=160, bbox_inches="tight")
plt.close(fig)

print(OUTPUT)
print(f"Doğrulandı: {len(comparison)} ortak ay, {len(headers)} sütun")
