from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "outputs" / "autogluon" / "ip7_devir_orani" / "test_tahminleri_tum_modeller.csv"
OUTPUT = Path(__file__).with_name("model_tahmin_karsilastirmasi.xlsx")

MONTHS_TR = {
    1: "Ocak",
    2: "Şubat",
    3: "Mart",
    4: "Nisan",
    5: "Mayıs",
    6: "Haziran",
    7: "Temmuz",
    8: "Ağustos",
    9: "Eylül",
    10: "Ekim",
    11: "Kasım",
    12: "Aralık",
}


raw = pd.read_csv(SOURCE, parse_dates=["hedef_ay"])
selected_name = (SOURCE.parent / "secili_model.txt").read_text(encoding="utf-8").strip()

main_model = (
    raw.loc[raw["model"].eq(selected_name), ["hedef_ay", "y_true", "y_pred"]]
    .rename(columns={"y_pred": "ana_model_tahmini"})
)
last_year = (
    raw.loc[raw["model"].eq("SeasonalNaive"), ["hedef_ay", "y_pred"]]
    .rename(columns={"y_pred": "gecen_yil_ayni_ay"})
)
comparison = main_model.merge(last_year, on="hedef_ay", validate="one_to_one").sort_values("hedef_ay")

if len(comparison) != 6:
    raise ValueError(f"Beklenen 6 test ayı yerine {len(comparison)} satır bulundu.")

wb = Workbook()
ws = wb.active
ws.title = "Tahmin Karşılaştırması"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:E1")
ws["A1"] = "Ana Model ve Geçen Yılın Aynı Ayı Karşılaştırması"
ws["A1"].font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="17365D")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

headers = [
    "Tahmin edilen ay",
    "Yıl",
    "Ana modelin tahmini",
    "Gerçek tahmin",
    "Geçen yılın o ayki değeri",
]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F75B5")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, row in enumerate(comparison.itertuples(index=False), start=4):
    ws.cell(row=row_idx, column=1, value=MONTHS_TR[row.hedef_ay.month])
    ws.cell(row=row_idx, column=2, value=int(row.hedef_ay.year))
    ws.cell(row=row_idx, column=3, value=float(row.ana_model_tahmini))
    ws.cell(row=row_idx, column=4, value=float(row.y_true))
    ws.cell(row=row_idx, column=5, value=float(row.gecen_yil_ayni_ay))

    for col in range(1, 6):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(name="Aptos", size=11)
        cell.alignment = Alignment(
            horizontal="left" if col == 1 else "right",
            vertical="center",
        )
    for col in range(3, 6):
        ws.cell(row=row_idx, column=col).number_format = "0.000%"

table = Table(displayName="TahminKarsilastirmaTablosu", ref=f"A3:E{3 + len(comparison)}")
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws.add_table(table)

thin_gray = Side(style="thin", color="D9E2F3")
ws["A3"].border = Border(left=thin_gray, top=thin_gray, bottom=thin_gray)
ws[f"E{3 + len(comparison)}"].border = Border(right=thin_gray, bottom=thin_gray)

ws.column_dimensions["A"].width = 21
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 23
ws.column_dimensions["D"].width = 19
ws.column_dimensions["E"].width = 29
ws.row_dimensions[3].height = 34
for row_idx in range(4, 4 + len(comparison)):
    ws.row_dimensions[row_idx].height = 21

note_row = 5 + len(comparison)
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
ws.cell(note_row, 1, "Not: Oranlar hesaplamaya uygun ham değer olarak saklanmış, hücrelerde yüzde biçiminde gösterilmiştir.")
ws.cell(note_row, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[note_row].height = 28

ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:E{3 + len(comparison)}"
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.print_title_rows = "1:3"
ws.print_area = f"A1:E{note_row}"

wb.save(OUTPUT)

# Dosyayı yeniden açarak içerik, satır sayısı ve sayı türlerini doğrula.
check = load_workbook(OUTPUT, data_only=False)
check_ws = check["Tahmin Karşılaştırması"]
assert [check_ws.cell(3, c).value for c in range(1, 6)] == headers
assert check_ws.max_row == note_row
assert all(isinstance(check_ws.cell(r, c).value, (int, float)) for r in range(4, 4 + len(comparison)) for c in range(2, 6))
assert check_ws["C4"].number_format == "0.000%"
assert check_ws.cell(3 + len(comparison), 5).number_format == "0.000%"

print(OUTPUT)
print(f"Doğrulandı: {len(comparison)} satır, {len(headers)} sütun")
